from __future__ import annotations

import csv
import inspect
import json
import os
import secrets
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Awaitable, Callable, Iterable, Tuple

from openpyxl import load_workbook
from telethon.errors import PhoneNumberUnoccupiedError, SessionPasswordNeededError

from src.accounts.connection_manager import ClientFactory
from src.accounts.session_crypto import ensure_encrypted
from src.config import AccountConfig, ProxyConfig

import logging
logger = logging.getLogger(__name__)

_REQUIRED_COLUMNS = {
    "APP_ID",
    "APP_HASH",
    "SDK",
    "DEVICE",
    "APP_VERSION",
    "LANG_CODE",
    "SYSTEM_LANG_CODE",
    "LANG_PACK",
    "TZ_OFFSET",
    "PERF_CAT",
}

_FIRST_NAMES = (
    "Alex",
    "Daniel",
    "David",
    "Emma",
    "James",
    "Julia",
    "Leo",
    "Maria",
    "Mia",
    "Nora",
    "Oliver",
    "Sofia",
)
_LAST_NAMES = (
    "Adams",
    "Brown",
    "Clark",
    "Davis",
    "Evans",
    "Hall",
    "Lewis",
    "Martin",
    "Miller",
    "Moore",
    "Taylor",
    "Wilson",
)


@dataclass(frozen=True)
class TelegramFingerprint:
    app_id: int
    app_hash: str
    sdk: int
    device: str
    app_version: str
    lang_code: str
    system_lang_code: str
    lang_pack: str
    tz_offset: int
    perf_cat: int

    def signature(self) -> Tuple[str, str]:
        """The (device, app_version) pair Telegram itself would see as this account's device."""
        return (self.device, self.app_version)

    @classmethod
    def from_mapping(cls, source: dict[str, Any], row_number: int) -> "TelegramFingerprint":
        values = {str(key).strip().upper(): value for key, value in source.items()}
        missing = sorted(
            column for column in _REQUIRED_COLUMNS
            if column not in values or values[column] is None or str(values[column]).strip() == ""
        )
        if missing:
            raise ValueError(
                f"Fingerprint row {row_number} is missing: {', '.join(missing)}"
            )

        try:
            app_id = int(values["APP_ID"])
            sdk = int(values["SDK"])
            tz_offset = int(values["TZ_OFFSET"])
            perf_cat = int(values["PERF_CAT"])
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"Fingerprint row {row_number} has an invalid numeric value"
            ) from exc

        app_hash = str(values["APP_HASH"]).strip()
        if app_id <= 0:
            raise ValueError(f"Fingerprint row {row_number} has an invalid APP_ID")
        if len(app_hash) != 32 or any(ch not in "0123456789abcdefABCDEF" for ch in app_hash):
            raise ValueError(f"Fingerprint row {row_number} has an invalid APP_HASH")

        return cls(
            app_id=app_id,
            app_hash=app_hash,
            sdk=sdk,
            device=str(values["DEVICE"]).strip(),
            app_version=str(values["APP_VERSION"]).strip(),
            lang_code=str(values["LANG_CODE"]).strip(),
            system_lang_code=str(values["SYSTEM_LANG_CODE"]).strip(),
            lang_pack=str(values["LANG_PACK"]).strip(),
            tz_offset=tz_offset,
            perf_cat=perf_cat,
        )


class FingerprintCatalog:
    def __init__(self, rows: list[TelegramFingerprint]) -> None:
        if not rows:
            raise ValueError("The Telegram fingerprint file contains no data rows")
        self._rows = rows

    @classmethod
    def load(cls, path: str | Path) -> "FingerprintCatalog":
        fingerprint_path = Path(path)
        if not fingerprint_path.is_file():
            raise FileNotFoundError(
                f"Telegram fingerprint file not found: {fingerprint_path}"
            )

        suffix = fingerprint_path.suffix.lower()
        if suffix == ".csv":
            with fingerprint_path.open("r", encoding="utf-8-sig", newline="") as stream:
                records = list(csv.DictReader(stream))
        elif suffix == ".json":
            with fingerprint_path.open("r", encoding="utf-8") as stream:
                payload = json.load(stream)
            records = payload if isinstance(payload, list) else payload.get("rows", [])
            if not isinstance(records, list):
                raise ValueError("Fingerprint JSON must be an array or contain a rows array")
        elif suffix in {".xlsx", ".xlsm"}:
            workbook = load_workbook(fingerprint_path, read_only=True, data_only=True)
            worksheet = workbook.active
            iterator = worksheet.iter_rows(values_only=True)
            headers = next(iterator, None)
            if not headers:
                raise ValueError("The Telegram fingerprint workbook is empty")
            records = [
                dict(zip(headers, row, strict=False))
                for row in iterator
                if any(value is not None and str(value).strip() for value in row)
            ]
            workbook.close()
        else:
            raise ValueError("Fingerprint file must be CSV, JSON, XLSX, or XLSM")

        if not records:
            raise ValueError("The Telegram fingerprint file contains no data rows")
        if not isinstance(records[0], dict):
            raise ValueError("Each Telegram fingerprint row must be an object")

        return cls(
            [
                TelegramFingerprint.from_mapping(record, index)
                for index, record in enumerate(records, start=2)
            ]
        )

    def choose(self, *, avoid: Iterable[Tuple[str, str]] = ()) -> TelegramFingerprint:
        """
        Picks a random fingerprint, preferring one not already in `avoid`
        (typically every currently-loaded account's (device, app_version)) so
        two accounts don't end up presenting as the identical device to
        Telegram -- an easy correlation signal, especially if they also share
        a proxy or the operator's own IP. Falls back to the full catalog
        (duplicates allowed) once every distinct signature is already in use.
        """
        avoid_set = set(avoid)
        candidates = [row for row in self._rows if row.signature() not in avoid_set]
        if not candidates:
            logger.warning(
                "FingerprintCatalog: all %d distinct device profiles are already in "
                "use by the pool; assigning a repeat at random.",
                len({row.signature() for row in self._rows}),
            )
            candidates = self._rows
        return secrets.choice(candidates)


@dataclass
class PendingTelegramLogin:
    phone: str
    account: AccountConfig
    fingerprint: TelegramFingerprint
    client: Any
    temp_dir: Path
    phone_code_hash: str
    first_name: str
    last_name: str
    created_new_account: bool = False
    authorized: bool = False


@dataclass(frozen=True)
class SavedTelegramSession:
    account: AccountConfig
    session_file: str
    first_name: str
    last_name: str
    created_new_account: bool


class TelegramAuthenticator:
    """Creates authorized Telethon sessions from Hero SMS phone/code pairs."""

    def __init__(
        self,
        *,
        fingerprint_file: str,
        accounts_dir: str,
        encryption_key: bytes | None = None,
        proxy_provider: Callable[
            [], ProxyConfig | None | Awaitable[ProxyConfig | None]
        ] | None = None,
        fingerprint_signatures_provider: Callable[[], Iterable[Tuple[str, str]]] | None = None,
    ) -> None:
        self._fingerprint_file = fingerprint_file
        self._accounts_dir = Path(accounts_dir)
        self._encryption_key = encryption_key
        self._catalog: FingerprintCatalog | None = None
        self._proxy_provider = proxy_provider
        self._fingerprint_signatures_provider = fingerprint_signatures_provider

    def validate_ready(self) -> None:
        try:
            self._load_catalog()
        except FileNotFoundError as exc:
            raise ValueError(str(exc)) from exc
        self._accounts_dir.mkdir(parents=True, exist_ok=True)

    async def begin(self, phone: str) -> PendingTelegramLogin:
        logger.debug("TelegramAuthenticator: beginning login for phone %s", phone)
        self.validate_ready()
        normalized = self._normalize_phone(phone)
        target_base = self._accounts_dir / normalized.lstrip("+")
        if (
            target_base.with_suffix(".session").exists()
            or Path(f"{target_base}.session.enc").exists()
            or target_base.with_suffix(".json").exists()
        ):
            raise FileExistsError(f"An account session already exists for {normalized}")

        avoid = (
            self._fingerprint_signatures_provider()
            if self._fingerprint_signatures_provider is not None
            else ()
        )
        fingerprint = self._load_catalog().choose(avoid=avoid)
        logger.debug("TelegramAuthenticator: using fingerprint %s", fingerprint)
        temp_dir = Path(
            tempfile.mkdtemp(prefix=".telegram-login-", dir=self._accounts_dir)
        )
        proxy = await self._next_proxy()
        account = self._account_from_fingerprint(
            normalized, fingerprint, session_dir=str(temp_dir), proxy=proxy
        )
        logger.debug("TelegramAuthenticator: building client for account %s", account)
        client = ClientFactory.build(account)
        logger.debug("TelegramAuthenticator: client built for account %s", account)
        first_name = secrets.choice(_FIRST_NAMES)
        last_name = secrets.choice(_LAST_NAMES)
        try:
            await client.connect()
            sent_code = await client.send_code_request(normalized)
            
            return PendingTelegramLogin(
                phone=normalized,
                account=account,
                fingerprint=fingerprint,
                client=client,
                temp_dir=temp_dir,
                phone_code_hash=sent_code.phone_code_hash,
                first_name=first_name,
                last_name=last_name,
            )
        except Exception:
            await self._disconnect(client)
            shutil.rmtree(temp_dir, ignore_errors=True)
            raise

    async def submit_code(self, login: PendingTelegramLogin, code: str) -> bool:
        """Returns False when a cloud 2FA password is still required."""
        try:
            await login.client.sign_in(
                phone=login.phone,
                code=code.strip(),
                phone_code_hash=login.phone_code_hash,
            )
        except PhoneNumberUnoccupiedError:
            await login.client.sign_up(
                code=code.strip(),
                first_name=login.first_name,
                last_name=login.last_name,
                phone=login.phone,
                phone_code_hash=login.phone_code_hash,
            )
            login.created_new_account = True
        except SessionPasswordNeededError:
            return False

        login.authorized = await login.client.is_user_authorized()
        if not login.authorized:
            raise PermissionError("Telegram did not authorize the new session")
        return True

    async def submit_password(
        self, login: PendingTelegramLogin, password: str
    ) -> None:
        if not password:
            raise ValueError("Telegram 2-step verification password is required")
        await login.client.sign_in(password=password)
        login.authorized = await login.client.is_user_authorized()
        if not login.authorized:
            raise PermissionError("Telegram did not authorize the new session")

    async def save(self, login: PendingTelegramLogin) -> SavedTelegramSession:
        if not login.authorized:
            raise PermissionError("Cannot save an unauthorized Telegram session")

        await self._disconnect(login.client)
        normalized_name = login.phone.lstrip("+")
        source = Path(f"{login.account.session_path}.session")
        target = self._accounts_dir / f"{normalized_name}.session"
        target_json = self._accounts_dir / f"{normalized_name}.json"
        if not source.exists():
            raise FileNotFoundError("Telethon did not create a session file")
        if target.exists() or Path(f"{target}.enc").exists() or target_json.exists():
            raise FileExistsError(f"An account session already exists for {login.phone}")

        os.replace(source, target)
        metadata = {
            "app_id": login.fingerprint.app_id,
            "app_hash": login.fingerprint.app_hash,
            "phone": login.phone,
            "device": login.fingerprint.device,
            "system": f"SDK {login.fingerprint.sdk}",
            "app_version": login.fingerprint.app_version,
            "lang_code": login.fingerprint.lang_code,
            "system_lang_code": login.fingerprint.system_lang_code,
            "lang_pack": login.fingerprint.lang_pack,
            "sdk": login.fingerprint.sdk,
            "tz_offset": login.fingerprint.tz_offset,
            "perf_cat": login.fingerprint.perf_cat,
            "first_name": login.first_name,
            "last_name": login.last_name,
            "created_new_account": login.created_new_account,
        }
        metadata_tmp = target_json.with_suffix(f".json.{secrets.token_hex(4)}.tmp")
        try:
            with metadata_tmp.open("w", encoding="utf-8") as stream:
                json.dump(metadata, stream, ensure_ascii=False, indent=2)
            os.replace(metadata_tmp, target_json)
        except Exception:
            target.unlink(missing_ok=True)
            metadata_tmp.unlink(missing_ok=True)
            raise
        finally:
            shutil.rmtree(login.temp_dir, ignore_errors=True)

        account = self._account_from_fingerprint(
            login.phone,
            login.fingerprint,
            session_dir=str(self._accounts_dir),
            proxy=login.account.proxy,
        )
        session_file = str(target)
        if self._encryption_key is not None:
            ensure_encrypted(account.session_path, self._encryption_key)
            session_file = f"{target}.enc"

        return SavedTelegramSession(
            account=account,
            session_file=session_file,
            first_name=login.first_name,
            last_name=login.last_name,
            created_new_account=login.created_new_account,
        )

    async def discard(self, login: PendingTelegramLogin) -> None:
        await self._disconnect(login.client)
        shutil.rmtree(login.temp_dir, ignore_errors=True)

    def _load_catalog(self) -> FingerprintCatalog:
        if self._catalog is None:
            self._catalog = FingerprintCatalog.load(self._fingerprint_file)
        return self._catalog

    async def _next_proxy(self) -> ProxyConfig | None:
        if self._proxy_provider is None:
            return None
        proxy = self._proxy_provider()
        if inspect.isawaitable(proxy):
            proxy = await proxy
        return proxy

    @staticmethod
    def _normalize_phone(phone: str) -> str:
        digits = "".join(character for character in phone if character.isdigit())
        if not digits:
            raise ValueError("Hero SMS returned an invalid phone number")
        return f"+{digits}"

    @staticmethod
    def _account_from_fingerprint(
        phone: str,
        fingerprint: TelegramFingerprint,
        *,
        session_dir: str,
        proxy: ProxyConfig | None = None,
    ) -> AccountConfig:
        return AccountConfig(
            api_id=fingerprint.app_id,
            api_hash=fingerprint.app_hash,
            phone=phone,
            proxy=proxy,
            device_model=fingerprint.device,
            system_version=f"SDK {fingerprint.sdk}",
            app_version=fingerprint.app_version,
            lang_code=fingerprint.lang_code,
            system_lang_code=fingerprint.system_lang_code,
            lang_pack=fingerprint.lang_pack,
            sdk=fingerprint.sdk,
            tz_offset=fingerprint.tz_offset,
            perf_cat=fingerprint.perf_cat,
            session_dir=session_dir,
        )

    @staticmethod
    async def _disconnect(client: Any) -> None:
        try:
            if client.is_connected():
                await client.disconnect()
        except Exception:
            pass
